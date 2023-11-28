import pygame
import random

from cv2 import VideoCapture
from cv2 import imshow
from cv2 import imwrite

from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

from email.message import EmailMessage
import ssl
import smtplib

import cv2
import face_recognition

# Initialize the game
pygame.init()
screen = pygame.display.set_mode((1600, 900), pygame.RESIZABLE)
pygame.display.set_caption("Et Bi 99 - NHA CAI DEN TU CHAU A")
clock = pygame.time.Clock()

# Show coin
collected_coin = 0
font = pygame.font.Font(None, 50)
text = font.render("COIN: ", True,  "#f06e4b")
text_rect = text.get_rect(center = (100, 50))

# To switch between scence
# 0: Main menu
# 1: Create account
# 2: Login
# 3: Mini game 
game_state = 0

# Minigame Player class
class Player():
    def __init__(self, x, y, image, scale):
        self.image = pygame.image.load(image)
        self.image = pygame.transform.scale(self.image, scale)
        self.rect = self.image.get_rect()
        self.rect.topleft = (x, y)

    def draw(self):
        screen.blit(self.image, (self.rect.x, self.rect.y))

    def isCollide(self, x, y):
        return self.rect.collidepoint(x, y)
    
    def dir_animate(self, value):
        if (value >= 0):
            self.image = pygame.transform.flip(self.image, 1, 0)

# Text Class
class Txt():
    def __init__(self, x, y, content, color, isBorder = False, isClickable = False):
        font = pygame.font.Font('BDLifelessGrotesk-Bold.otf', 30)
        self.text = font.render(content, True, color)
        self.rect = self.text.get_rect()
        self.rect.topleft = (x, y)
        self.isBorder = isBorder
        self.isClickable = isClickable

    def render(self):
        if self.isBorder:
            rect_color = "#726f6f"
            rect_position = self.rect.topleft
            rect_size = (self.text.get_width() + 5, self.text.get_height() + 5)
            
            pygame.draw.rect(screen, rect_color, (rect_position, rect_size))
        
        if self.isClickable:
            pos = pygame.mouse.get_pos()
            if self.rect.collidepoint(pos):
                self.text = pygame.transform.scale(self.text, (self.text.get_width() * 1.1, self.text.get_height() * 1.1))

        screen.blit(self.text, self.rect)

    def isClick(self):
        pos = pygame.mouse.get_pos()

        if self.rect.collidepoint(pos) and pygame.mouse.get_pressed()[0] == 1:
            return True
        else:
            return False  

# Image Class
class Img():
    def __init__(self, x, y, image, scale):
        self.image = pygame.image.load(image)
        self.image = pygame.transform.scale(self.image, scale)
        self.rect = self.image.get_rect()
        self.rect.topleft = (x, y)

    def draw(self):
        screen.blit(self.image, (self.rect.x, self.rect.y))

# Item Class
class Item():
    def __init__(self, x, y, image, scale, value):
        self.image = pygame.image.load(image)
        self.image = pygame.transform.scale(self.image, scale)
        self.rect = self.image.get_rect()
        self.rect.topleft = (x, y)
        self.value = value

    def draw(self):
        pos = pygame.mouse.get_pos()
        if self.rect.collidepoint(pos):
            self.image = pygame.transform.scale(self.image, (self.image.get_width() * 1.1, self.image.get_height() * 1.1))
        screen.blit(self.image, (self.rect.x, self.rect.y))

    
    def isClick(self):
        pos = pygame.mouse.get_pos()
        if self.rect.collidepoint(pos) and pygame.mouse.get_pressed()[0] == 1:
            return True
        else:
            return False  

    def getValue(self):
        return self.value

# Button Class
class Button():
    def __init__(self, x, y, image, scale):
        self.image = pygame.image.load(image)
        self.image = pygame.transform.scale(self.image, scale)
        self.rect = self.image.get_rect()
        self.rect.topleft = (x, y)
        self.scale = scale

    def isClick(self):
        pos = pygame.mouse.get_pos()
        if self.rect.collidepoint(pos) and pygame.mouse.get_pressed()[0] == 1:
            return True
        else:
            return False  

    def draw(self):
        pos = pygame.mouse.get_pos()

        if self.rect.collidepoint(pos):
            self.image = pygame.transform.scale(self.image, (self.image.get_width() * 1.1, self.image.get_height() * 1.1))

        screen.blit(self.image, (self.rect.x, self.rect.y))

# Send mail to player after create account
def sendMail(sender, password, receiver, subject, body):
    mail = EmailMessage()
    mail['From'] = sender
    mail['To'] = receiver
    mail['Subject'] = subject
    mail.set_content(body)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
        smtp.login(sender, password)
        smtp.sendmail(sender, receiver, mail.as_string())

pygame.mixer.init()
pygame.mixer.music.load("./Asset/BG-Music-2.mp3")
pygame.mixer.music.play(-1, 0, 2000)

# Main background Grass moving animation
bg_1_x = 0
# Random mouse position in minigame
mg_mouse = (random.random() * 1500, 720)
# Game tick to calculate time in minigame
mg_tick = 0
# Cloud pos x to animate in minigame
bg_sun_x = 0

x_add = 0
y_add = 0

name = ""
email = ""
password = ""

info = []

log_name = ""
log_password = ""
isLogin = False

player = ""
coin = 0

pos = 1

LANG = 1

prePlayItems = set()
# Language for game
workbook = openpyxl.load_workbook("languages.xlsx")
sheet = workbook.active
rows = []
for row in sheet.iter_rows(values_only=True):
    rows.append(list(row))

lang = dict()
for row in rows:
    lang[row[0]] = row[LANG]
#--------------------------------

while True:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            pygame.quit()
            exit

        if event.type == pygame.KEYDOWN and game_state == 1:
            if event.key == pygame.K_BACKSPACE: 
                name = name[:-1] 
            else: 
                name += event.unicode

        if event.type == pygame.KEYDOWN and game_state == 11:
            if event.key == pygame.K_BACKSPACE: 
                email = email[:-1] 
            else: 
                email += event.unicode
    
        if event.type == pygame.KEYDOWN and game_state == 12:
            if event.key == pygame.K_BACKSPACE: 
                password = password[:-1] 
            else: 
                password += event.unicode

        if event.type == pygame.KEYDOWN and game_state == 2:
            if event.key == pygame.K_BACKSPACE: 
                log_name = log_name[:-1] 
            else: 
                log_name += event.unicode
        
        if event.type == pygame.KEYDOWN and game_state == 22:
            if event.key == pygame.K_BACKSPACE: 
                log_password = log_password[:-1] 
            else: 
                log_password += event.unicode

    w, h = pygame.display.get_surface().get_size()

    # Load main menu
    if game_state == 0:

        bg = Img(0, 0, "./Asset/BG.png", (w, h))
        bg.draw()
        bg_text = Img(w / 2 - 500, h / 5 + 15, "./Asset/" + lang['BG-TITLE'], (1000, 200))
        bg_text.draw()

        bg_1_x -= 2
        if bg_1_x <= -1000:
            bg_1_x = 0 
        bg_1 = Img(bg_1_x, h - 250, "./Asset/BG3.png", (4000, 250))
        bg_1.draw()

        btn_exit = Button(w / 2 + 325, 97, "./Asset/ExitGame.png", (50, 50))
        btn_exit.draw()
        if btn_exit.isClick():
            if isLogin:
                wb = load_workbook("player.xlsx")
                sheet = wb.active
                sheet['E' + chr(pos + 48)] = coin
                wb.save("player.xlsx")
            exit()

        btn_create_account = Button(w / 2 + 160, 98, "./Asset/" + lang['CREATE'], (150, 50))
        btn_create_account.draw()
        if btn_create_account.isClick():
            game_state = 1

            pygame.mixer.music.stop()
            pygame.mixer.music.unload()

            pygame.mixer.music.load("./Asset/BG-Music-1.mp3")
            pygame.mixer.music.play(-1, 0, 2000)

        btn_login = Button(w / 2, 98, "./Asset/" + lang['LOGIN'], (150, 50))
        btn_login.draw()
        if btn_login.isClick():
            game_state = 2

            pygame.mixer.music.stop()
            pygame.mixer.music.unload()

            pygame.mixer.music.load("./Asset/BG-Music-1.mp3")
            pygame.mixer.music.play(-1, 0, 2000)

        text = Txt(w / 2 - 300, 100, lang['PLAYER'] + player, "#f06e4b", True)
        text.render()

        text = Txt(w / 2 - 300, 150, lang['COIN'] + str(coin), "#f06e4b", True)
        text.render()

        btn_play = Button(w / 2 - 350, h / 2, "./Asset/" + lang['BTN-PLAY'], (150, 50))
        btn_play.draw()

        btn_minigame = Button(w / 2 - 350, h / 2 + 75 , "./Asset/" + lang['BTN-MINIGAME'], (150, 50))
        btn_minigame.draw()

        if (btn_minigame.isClick()):
            game_state = 3
            mg_tick = 620

            screen.fill("#96c3d7")
            pygame.mixer.music.stop()
            pygame.mixer.music.unload()

            pygame.mixer.music.load("./Minigame/MG-Music-1.mp3")
            pygame.mixer.music.play(-1, 0, 2000)
            collected_coin = 0

            pygame.mouse.set_cursor(pygame.SYSTEM_CURSOR_CROSSHAIR)


        text = Txt(w / 2 + 200, 550, lang['LANG'], "WHITE", True, True)
        text.render()

        btn_shop = Button(w / 2 - 350, h / 2 + 150 , "./Asset/" + lang['STORE'], (150, 50))
        btn_shop.draw()

        if btn_shop.isClick():
            game_state = 4

        if text.isClick():
            game_state = -1

    # Change language state
    elif game_state == -1:
        if LANG == 1:
            LANG = 2
        else:
            LANG = 1

        # Language for game
        workbook = openpyxl.load_workbook("languages.xlsx")
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))

        lang = dict()
        for row in rows:
            lang[row[0]] = row[LANG]
        # End Language for game

        game_state = 0

    # Create account state
    elif game_state == 1:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['ENTERYOURNAME'] + name, "WHITE")
        text.render()

        text = Txt(w / 2 + 250, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 11
       
    elif game_state == 11:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['ENTERYOUREMAIL'] + email, "WHITE")
        text.render()

        text = Txt(w / 2 + 350, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 12
    
    elif game_state == 12:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['ENTERYOURPASSWORD'] + password, "WHITE")
        text.render()

        text = Txt(w / 2 + 250, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 13


    elif game_state == 13:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['TAKEYOURPICTURE'], "WHITE")
        text.render()

        text = Txt(w / 2 + 350, 275, lang["TAKE"], "WHITE", True, True)
        text.render()

        if text.isClick():
            game_state = 0

            cam_port = 0
            cam = VideoCapture(cam_port)

            result, image = cam.read()
            if result:
                game_state = 14

                imshow("Taken Picture", image)
                imwrite("./player_img/" + name + ".png", image)

                info.append(name)
                info.append(email)
                info.append(name + ".png")
                info.append(password)
                info.append(20) # Give 20 coin to new player

                wb = load_workbook("player.xlsx")
                sheet = wb.active
                sheet.append(info)
                wb.save("player.xlsx")

                sendMail("ltloc05samsunggalaxyj3pro@gmail.com", "rodq twhi tmme gypg", info[1]
                , "Welcome to my game"
                , "Chuc mung ban " + info[0] + " tao tai khoan game ca cuoc thanh cong :)) From Nhom 9 - 23CTT1 - NMCNTT - HCMUS with love")

    elif game_state == 14:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['CREATESUCCESSFULLY'], "WHITE")
        text.render()

        text = Txt(w / 2 + 250, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 0

    elif game_state == 2:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['ENTERYOURNAME'] + log_name, "WHITE")
        text.render()

        text = Txt(w / 2 + 250, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 21 

    elif game_state == 21:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['USERANDPASSWORD'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 22

        text = Txt(w / 2 - 450, 250, lang['FACERECOGNITION'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 25

    elif game_state == 22:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 600, 200, lang["ENTERYOURPASSWORD"] + log_password, "WHITE")
        text.render()

        text = Txt(w / 2 + 250, 200, lang['NEXT'], "WHITE", True, True)
        text.render()

        if text.isClick():
            workbook = openpyxl.load_workbook("player.xlsx")
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row))
            
            game_state = 24

            i = 0
            for row in rows:
                i += 1
                if (row[0] == log_name and row[3] == log_password):
                    game_state = 23

                    isLogin = True
                    player = log_name
                    coin = row[4]
                    pos = i
                    break
            workbook.close()
    
    elif game_state == 23:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['LOGINSUCCESSFULLY'], "WHITE")
        text.render()

        text = Txt(w / 2 + 100, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 0 

    elif game_state == 24:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, lang['LOGINFAIL'], "WHITE")
        text.render()

        text = Txt(w / 2 + 100, 200, lang['NEXT'], "WHITE", True, True)
        text.render()
        if text.isClick():
            game_state = 0 

    elif game_state == 25:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 450, 200, "CHECKING YOUR FACE", "WHITE")
        text.render()

        # Load player image path to check
        known_image_path = "./player_img/" + log_name + ".png"
        known_image = face_recognition.load_image_file(known_image_path)
        known_face_encoding = face_recognition.face_encodings(known_image)[0]

        # Open the video capture
        cam_port = 0
        cam = cv2.VideoCapture(cam_port)  # Use 0 for default camera

        founded = False
        while True:
            result, image = cam.read()

            face_locations = face_recognition.face_locations(image)
            face_encodings = face_recognition.face_encodings(image, face_locations)

            # Loop through each face found in the frame
            for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                # Check if the face matches the known face
                matches = face_recognition.compare_faces([known_face_encoding], face_encoding)

                name = "Cannot verified"
                if matches[0]:
                    name = "Hi " + log_name +" welcome back to game. Please close this window"


                    workbook = openpyxl.load_workbook("player.xlsx")
                    sheet = workbook.active
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        rows.append(list(row))

                    game_state = 23

                    i = 0
                    for row in rows:
                        i += 1
                        if row[0] == log_name:
                            isLogin = True
                            player = log_name
                            coin = row[4]
                            pos = i

                    founded = True
                # Draw rectangle around the face
                cv2.rectangle(image, (left, top), (right, bottom), (0, 255, 0), 2)

                # Display the name of the person
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(image, name, (left + 6, bottom - 6), font, 0.5, (255, 255, 255), 1)

            # Display the resulting frame
            cv2.imshow('Checking face', image)
            
            if founded:
                break

    # Minigame
    elif game_state == 3:
        mg_tick -= 1

        # change mouse pos after 0.5s
        if mg_tick % 15 == 0:
            x_add = random.random() * 40 - 20
            y_add = random.random() * 40 - 20
        mg_mouse_x = mg_mouse[0]
        mg_mouse_x += x_add
        # when mouse go to screen border
        if mg_mouse_x <= 0:
            mg_mouse_x = w
        if mg_mouse_x >= w:
            mg_mouse_x = 0
        mg_mouse_y = mg_mouse[1]
        mg_mouse_y += y_add
        # when mouse go to screen border
        if mg_mouse_y <= h - 400:
            mg_mouse_y = h
        if mg_mouse_y >= h:
            mg_mouse_y = h - 400
        mg_mouse = (mg_mouse_x, mg_mouse_y)

        # time went off
        if (mg_tick == 0):
            game_state = 0
            pygame.mixer.Sound("./Minigame/MG-Win.mp3").play()

            # plus gameplay coin to player's coin
            coin += collected_coin

        screen.fill("#96c3d7")
        text = Txt(300, 100, lang['COLLECTEDCOIN'] + str(collected_coin), "#f06e4b", True, True)
        text.render()

        text = Txt(300, 50, lang['TIME'] + str(int(mg_tick / 30)), "#f06e4b", True, True)
        text.render()

        bg_grass = Img(0, h - 1000, "./Minigame/MG-Grass.png", (5000, 1000))
        bg_grass.draw()

        bg_sun = Img(300, 200, "./Minigame/MG-Sun.png", (100, 100))
        bg_sun.draw()

        bg_sun_x += 1
        if bg_sun_x >= h:
            bg_sun_x = 0
        bg_sun = Img(bg_sun_x, 300, "./Minigame/MG-Cloud.png", (1600, 100))
        bg_sun.draw()
        
        bg_1 = Player(mg_mouse[0], mg_mouse[1], "./Minigame/MG-Mouse.png", (250, 250))
        bg_1.dir_animate(x_add)
        bg_1.draw()

        mouse_input = pygame.mouse.get_pos()

        mg_hammer = Img(mouse_input[0] - 50, mouse_input[1] - 50, "./Minigame/MG-Hammer.png", (100, 100))
        mg_hammer.draw()

        if bg_1.isCollide(mouse_input[0], mouse_input[1]):
            if pygame.mouse.get_pressed()[0]:
                collected_coin += 1
                pygame.mixer.Sound("./Minigame/MG-Coin.mp3").play()

                mg_mouse = (random.random() * (w - 200), h - 500 + random.random() * 300)
    
    elif game_state == 4:
        screen.fill("#96c3d7")

        text = Txt(w / 2 - 300, 100, lang['PLAYER'] + player, "#f06e4b", True)
        text.render()

        text = Txt(w / 2 - 300, 150, lang['COIN'] + str(coin), "#f06e4b", True)
        text.render()

        text = Txt(w / 2 - 300, 200, lang['BUYED'] + str(len(prePlayItems)), "WHITE", True)
        text.render()

        if lang['ITEM-1'] not in prePlayItems:
            item_1 = Item(w / 2 - 300, 250, "./Asset/" + lang['ITEM-1'], (300, 100), 10)
            item_1.draw()
            if item_1.isClick():
                value = item_1.getValue()
                if coin >= value:
                    coin -= value
                    prePlayItems.add(lang['ITEM-1'])

        if lang['ITEM-2'] not in prePlayItems:
            item_2 = Item(w / 2 - 300, 350, "./Asset/" + lang['ITEM-2'], (300, 100), 10)
            item_2.draw()
            if item_2.isClick():
                value = item_2.getValue()
                if coin >= value:
                    coin -= value
                    prePlayItems.add(lang['ITEM-2'])

        if lang['ITEM-3'] not in prePlayItems:
            item_2 = Item(w / 2 - 300, 450, "./Asset/" + lang['ITEM-3'], (300, 100), 10)
            item_2.draw()
            if item_2.isClick():
                value = item_2.getValue()
                if coin >= value:
                    coin -= value
                    prePlayItems.add(lang['ITEM-3'])

        btn_exit = Button(w / 2 + 300, h / 2 + 100, "./Asset/ExitGame.png", (50, 50))
        btn_exit.draw()
        if btn_exit.isClick():
            game_state = 0
            
    pygame.display.update()
    clock.tick(30)